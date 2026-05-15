"""
backend/ccr.py
--------------
Completeness, Consistency, and Reasonableness (CCR) checks.

check_completeness       - mandatory tabs present + zero-emergence flags
check_consistency        - subtotal integrity + optional working-file cross-check
check_reasonableness     - large YoY swings and statistical outliers
check_working_file       - cross-check FCA values against uploaded working papers
"""

import pandas as pd
import numpy as np
import openpyxl
from rapidfuzz import fuzz, process
from typing import Optional


# ── Completeness ───────────────────────────────────────────────────────────

def check_completeness(cy_parsed: dict, template_tabs: list) -> list:
    """Flags mandatory tabs missing from the CY file."""
    findings = []
    for tab in template_tabs:
        if tab not in cy_parsed:
            findings.append({
                "tab":      tab,
                "check":    "Completeness",
                "finding":  f"Tab '{tab}' is missing from the uploaded CY file.",
                "severity": "High",
            })
    return findings


def check_zero_emergence(aligned_df: pd.DataFrame, tab_name: str) -> list:
    """
    Flags two critical scenarios:
      - DISAPPEARED: PY had a non-zero value, CY is now zero or absent.
        (Why would a risk/asset vanish entirely?)
      - EMERGED: CY has a non-zero value, PY was zero or absent.
        (A new material item has appeared — needs explanation.)

    Only flags rows above a minimum absolute value threshold to avoid noise.
    """
    findings = []
    MIN_VALUE = 1_000   # ignore sub-1k items (rounding noise)

    df = aligned_df.copy()
    df["cy_value"] = pd.to_numeric(df["cy_value"], errors="coerce")
    df["py_value"] = pd.to_numeric(df["py_value"], errors="coerce")

    for _, row in df.iterrows():
        cy = row.get("cy_value")
        py = row.get("py_value")
        desc = row.get("description", row.get("cy_code", "Unknown"))

        # DISAPPEARED: had meaningful PY value, now zero/null
        if (pd.notna(py) and abs(py) >= MIN_VALUE
                and (pd.isna(cy) or abs(cy) < 1)):
            findings.append({
                "tab":      tab_name,
                "check":    "Completeness",
                "finding":  (
                    f"DISAPPEARED: '{desc}' had PY value of {py:,.0f} "
                    f"but is zero or absent in CY. "
                    f"Confirm intentional removal or reclassification."
                ),
                "severity": "High",
            })

        # EMERGED: new non-zero CY value, was zero/null in PY
        elif (pd.notna(cy) and abs(cy) >= MIN_VALUE
              and (pd.isna(py) or abs(py) < 1)):
            findings.append({
                "tab":      tab_name,
                "check":    "Completeness",
                "finding":  (
                    f"EMERGED: '{desc}' is new in CY with value {cy:,.0f} "
                    f"(was zero or absent in PY). "
                    f"Confirm this is a genuine new item."
                ),
                "severity": "Medium",
            })

    return findings


# ── Consistency ────────────────────────────────────────────────────────────

def check_consistency(df: pd.DataFrame, tab_name: str) -> list:
    """Flags subtotal rows where the declared total doesn't match sum of children."""
    findings = []

    cy_codes = df["cy_code"].dropna()
    total_rows = df[
        cy_codes.str.endswith((".T", ".0", "U.", "L.", "I."), na=False) |
        cy_codes.str.contains(r"\.U$|\.L$|\.I\.$", na=False, regex=True)
    ]

    # Also catch rows where description contains "total" or "summation"
    if "description" in df.columns:
        desc_totals = df[
            df["description"].str.lower().str.contains(
                "summation|total assets|total liabilities|total equity",
                na=False
            )
        ]
        total_rows = pd.concat([total_rows, desc_totals]).drop_duplicates()

    for _, total_row in total_rows.iterrows():
        code = total_row.get("cy_code", "")
        if not code:
            continue
        prefix = code.rsplit(".", 1)[0]
        children = df[
            df["cy_code"].str.startswith(prefix, na=False)
            & ~df["cy_code"].isin([code])
            & ~df["cy_code"].str.contains("summation|total", na=False, case=False)
        ]

        if len(children) < 2:
            continue

        try:
            expected = pd.to_numeric(children["cy_value"], errors="coerce").sum()
            actual   = float(total_row["cy_value"])
            if abs(expected - actual) > max(1, abs(actual) * 0.001):
                findings.append({
                    "tab":      tab_name,
                    "check":    "Consistency",
                    "finding":  (
                        f"Subtotal {code} = {actual:,.0f} "
                        f"but sum of {len(children)} children = {expected:,.0f}. "
                        f"Discrepancy: {actual - expected:,.0f}"
                    ),
                    "severity": "High",
                })
        except (TypeError, ValueError):
            continue

    return findings


def check_working_file(fca_data: dict, working_filepath: str,
                       tolerance_pct: float = 1.0) -> list:
    """
    Cross-checks key numeric values in a working paper Excel file against
    the corresponding values in the FCA return.

    Returns a list of discrepancy findings.
    tolerance_pct: percentage difference treated as acceptable (default 1%).
    """
    findings = []

    try:
        wb = openpyxl.load_workbook(working_filepath, data_only=True)
    except Exception as e:
        return [{
            "tab": "Working File",
            "check": "Consistency",
            "finding": f"Could not open working file: {e}",
            "severity": "High",
        }]

    # Build a flat dict of { label: value } from working file
    # Search all sheets for rows with a text label and adjacent number
    working_values: dict[str, float] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row: continue
            for ci in range(len(row) - 1):
                label = row[ci]
                value = row[ci + 1]
                if label is None or value is None: continue
                label_str = str(label).strip()
                if len(label_str) < 3: continue
                try:
                    num = float(value)
                    if abs(num) > 100:   # ignore small/zero values
                        working_values[label_str.lower()] = num
                except (TypeError, ValueError):
                    continue

    if not working_values:
        return [{
            "tab": "Working File",
            "check": "Consistency",
            "finding": "No numeric data found in working file.",
            "severity": "Medium",
        }]

    # Build flat dict from FCA data
    fca_values: dict[str, tuple[str, str, float]] = {}  # label -> (tab, code, value)
    for tab_name, df in fca_data.items():
        for code, row in df.iterrows():
            desc = str(row.get("description", code)).strip().lower()
            val  = row.get("value")
            try:
                num = float(val)
                if abs(num) > 100:
                    fca_values[desc] = (tab_name, str(code), num)
            except (TypeError, ValueError):
                continue

    fca_labels = list(fca_values.keys())

    # Fuzzy-match working file labels against FCA labels
    matched = 0
    for wk_label, wk_val in working_values.items():
        result = process.extractOne(wk_label, fca_labels,
                                    scorer=fuzz.token_sort_ratio)
        if not result or result[1] < 80:
            continue

        best_label = result[0]
        tab_name, code, fca_val = fca_values[best_label]

        if fca_val == 0:
            continue

        diff_pct = abs((wk_val - fca_val) / fca_val) * 100
        matched += 1

        if diff_pct > tolerance_pct:
            findings.append({
                "tab":      tab_name,
                "check":    "Consistency",
                "finding":  (
                    f"WORKING FILE MISMATCH: '{best_label}' ({code})\n"
                    f"  FCA return: {fca_val:,.0f}  |  "
                    f"Working file: {wk_val:,.0f}  |  "
                    f"Difference: {diff_pct:.2f}%"
                ),
                "severity": "High" if diff_pct > 5 else "Medium",
            })

    if matched == 0:
        findings.append({
            "tab": "Working File",
            "check": "Consistency",
            "finding": "No matching items found between working file and FCA return. "
                       "Check that the working file uses similar account labels.",
            "severity": "Medium",
        })

    return findings


# ── Reasonableness ─────────────────────────────────────────────────────────

def check_reasonableness(df: pd.DataFrame, tab_name: str) -> list:
    """Flags statistically unusual YoY movements using z-score and % threshold."""
    findings = []

    PCT_THRESHOLD = 50.0   # flag if |% change| > 50%
    ZSCORE_HIGH   = 3.0
    ZSCORE_MEDIUM = 2.0

    high_movers = df[
        (df["pct_change"].abs() > PCT_THRESHOLD) |
        (df["z_score"].abs() > ZSCORE_MEDIUM)
    ].dropna(subset=["pct_change"])

    for _, row in high_movers.iterrows():
        z   = float(row["z_score"]) if pd.notna(row.get("z_score")) else 0.0
        pct = float(row["pct_change"])
        cy  = row.get("cy_value", 0) or 0
        py  = row.get("py_value", 0) or 0
        desc = row.get("description", row.get("cy_code", "Unknown"))

        severity = "High" if abs(z) > ZSCORE_HIGH or abs(pct) > 100 else "Medium"

        findings.append({
            "tab":      tab_name,
            "check":    "Reasonableness",
            "finding":  (
                f"{desc}: {pct:+.1f}% YoY change "
                f"(z={z:.1f}\u03c3). "
                f"CY={cy:,.0f}, PY={py:,.0f}. "
                f"{'Statistically extreme — warrants investigation.' if abs(z) > ZSCORE_HIGH else 'Above materiality threshold.'}"
            ),
            "severity": severity,
        })

    return findings