"""
backend/ca1_parser.py
---------------------
Extracts HKRBC capital figures from CA.1 CA Summary and F.1 EBS.

CA.1 structure (confirmed from debug):
  - Row labels live in Col1 (index 1)  — NOT Col0
  - Numeric values live in Col2 (index 2)
  - Ratios stored as plain decimals e.g. 5.045 = 504.5%

F.1 EBS structure:
  - Row codes like "I.U", "II.L" live in Col0 (index 0)
  - Values in Col1 (index 1)
"""

import openpyxl, math, json, httpx, os
from dotenv import load_dotenv
from typing import Optional

load_dotenv()

BASE_URL = os.getenv("LLM_BASE_URL", "https://nova.deloitte.com.cn/del/v1")
API_KEY  = os.getenv("LLM_API_KEY", "")
MODEL    = os.getenv("LLM_MODEL", "Kimi-K2.5")


def _to_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        f = float(val)
        return None if (math.isnan(f) or math.isinf(f)) else f
    except (ValueError, TypeError):
        return None


def _find_ca1(rows: list, keywords: list) -> Optional[float]:
    """
    Search CA.1 rows where Col1 (index 1) contains any keyword.
    Return numeric value from Col2 (index 2).
    """
    kws = [k.lower().strip() for k in keywords]
    for row in rows:
        if not row or len(row) < 2 or row[1] is None:
            continue
        label = str(row[1]).strip().lower()
        if any(k in label for k in kws):
            val = _to_float(row[2]) if len(row) > 2 else None
            if val is not None:
                return val
    return None


def _find_f1(rows: list, prefix: str) -> Optional[float]:
    """
    Search F.1 EBS rows where Col0 (index 0) starts with prefix.
    Return numeric value from Col1 (index 1).
    """
    for row in rows:
        if not row or row[0] is None:
            continue
        if str(row[0]).strip().startswith(prefix):
            val = _to_float(row[1]) if len(row) > 1 else None
            if val is not None:
                return val
    return None


def _extract_single_file(filepath: str) -> dict:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}

    # ── CA.1 CA Summary ────────────────────────────────────────────────────
    ca1_ws = None
    for name in wb.sheetnames:
        n = name.upper()
        if "CA.1" in n and "CA.R" not in n and "CA.P" not in n:
            ca1_ws = wb[name]
            break

    if ca1_ws:
        rows = list(ca1_ws.iter_rows(values_only=True))

        result["eligible_capital"]    = _find_ca1(rows, ["total eligible capital"])
        result["tier1_capital"]       = _find_ca1(rows, ["total tier 1 eligible capital"])
        result["market_risk"]         = _find_ca1(rows, ["market risk"])
        result["life_insurance_risk"] = _find_ca1(rows, ["life insurance risk"])
        result["op_risk"]             = _find_ca1(rows, ["operational risk"])
        result["counterparty_risk"]   = _find_ca1(rows, ["counterparty default"])
        result["pca_before_op"]       = _find_ca1(rows, ["pca before operational risk",
                                                          "pca before ope"])
        # PCA — row 41 is plain "PCA", row 43 is "PCA (after specified…)"
        # Prefer "after specified" first, fall back to plain "PCA"
        pca = _find_ca1(rows, ["pca (after spe", "pca (after specified"])
        if pca is None:
            pca = _find_ca1(rows, ["pca (before lac", "pca (before"])
        if pca is None:
            # Plain "PCA" row — but avoid matching "PCA before..." or "PCA (..."
            for row in rows:
                if not row or len(row) < 2 or row[1] is None:
                    continue
                label = str(row[1]).strip().lower()
                if label == "pca":
                    v = _to_float(row[2]) if len(row) > 2 else None
                    if v is not None:
                        pca = v
                        break
        result["pca"] = pca

        mca = _find_ca1(rows, ["mca (after spe", "mca (after specified"])
        if mca is None:
            for row in rows:
                if not row or len(row) < 2 or row[1] is None:
                    continue
                label = str(row[1]).strip().lower()
                if label == "mca":
                    v = _to_float(row[2]) if len(row) > 2 else None
                    if v is not None:
                        mca = v
                        break
        result["mca"] = mca

        # Ratios: official template uses IFERROR formula cells — data_only returns None.
        # Compute directly from eligible capital / PCA / MCA values.
        ec    = result.get("eligible_capital")
        pca_v = result.get("pca")
        mca_v = result.get("mca")
        if ec and pca_v and pca_v > 0:
            result["pca_ratio"] = round(ec / pca_v * 100, 2)
        if ec and mca_v and mca_v > 0:
            result["mca_ratio"] = round(ec / mca_v * 100, 2)

        # Fallback: try reading pre-filled ratio cells (some insurers store values)
        if not result.get("pca_ratio") or not result.get("mca_ratio"):
            for row in rows:
                if not row or len(row) < 2 or row[1] is None:
                    continue
                label = str(row[1]).strip().lower()
                if "ratio of eligible capital" not in label:
                    continue
                # Try columns 2 and 3
                v = _to_float(row[2]) if len(row) > 2 else None
                if v is None:
                    v = _to_float(row[3]) if len(row) > 3 else None
                if v is None:
                    continue
                # Convert decimal (2.48) to percentage (248) if needed
                pct = v * 100 if v < 20 else v
                if "to mca" in label and not result.get("mca_ratio"):
                    result["mca_ratio"] = pct
                elif "to pca" in label and not result.get("pca_ratio"):
                    result["pca_ratio"] = pct

    # ── F.1 EBS ────────────────────────────────────────────────────────────
    f1_ws = None
    for name in wb.sheetnames:
        n = name.upper()
        if n.startswith("F.1") and "EBS" in n and "F.1B" not in n:
            f1_ws = wb[name]
            break

    if f1_ws:
        rows = list(f1_ws.iter_rows(values_only=True))
        result["total_assets"]          = _find_f1(rows, "I.U")
        result["total_liabilities"]     = _find_f1(rows, "II.L")
        result["moce"]                  = _find_f1(rows, "II.B4")
        result["inforce_reserves"]      = _find_f1(rows, "II.B3")
        result["total_equity"]          = _find_f1(rows, "III.I")
        # Insurance liabilities — search col 1 for text
        for row in rows:
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip().lower()
            if "total insurance liabilities" in label:
                v = _to_float(row[1]) if len(row) > 1 else None
                if v is not None:
                    result["insurance_liabilities"] = v
                    break

    ins  = result.get("insurance_liabilities") or 0
    moce = result.get("moce") or 0
    result["ce"] = max(0.0, ins - moce)

    return result


def _generate_capital_commentary(cy: dict, py: dict) -> str:
    def fmt(v):   return f"{v:,.0f}" if v is not None else "N/A"
    def pchg(a, b):
        if a and b and b != 0:
            return f"{((a-b)/abs(b))*100:+.1f}%"
        return "N/A"

    msg = f"""
HKRBC Capital Adequacy — YE25 (CY) vs YE24 (PY), HKD thousands

                         YE25           YE24        Change
Total Assets:         {fmt(cy.get('total_assets'))}  {fmt(py.get('total_assets'))}  {pchg(cy.get('total_assets'), py.get('total_assets'))}
Total Liabilities:    {fmt(cy.get('total_liabilities'))}  {fmt(py.get('total_liabilities'))}  {pchg(cy.get('total_liabilities'), py.get('total_liabilities'))}
Eligible Capital:     {fmt(cy.get('eligible_capital'))}  {fmt(py.get('eligible_capital'))}  {pchg(cy.get('eligible_capital'), py.get('eligible_capital'))}
PCA:                  {fmt(cy.get('pca'))}  {fmt(py.get('pca'))}  {pchg(cy.get('pca'), py.get('pca'))}
MCA:                  {fmt(cy.get('mca'))}  {fmt(py.get('mca'))}  {pchg(cy.get('mca'), py.get('mca'))}
PCA Coverage Ratio:   {f"{cy.get('pca_ratio'):.1f}%" if cy.get('pca_ratio') else 'N/A'}    {f"{py.get('pca_ratio'):.1f}%" if py.get('pca_ratio') else 'N/A'}
MCA Coverage Ratio:   {f"{cy.get('mca_ratio'):.1f}%" if cy.get('mca_ratio') else 'N/A'}    {f"{py.get('mca_ratio'):.1f}%" if py.get('mca_ratio') else 'N/A'}
Market Risk:          {fmt(cy.get('market_risk'))}  {fmt(py.get('market_risk'))}  {pchg(cy.get('market_risk'), py.get('market_risk'))}
Life Insurance Risk:  {fmt(cy.get('life_insurance_risk'))}  {fmt(py.get('life_insurance_risk'))}  {pchg(cy.get('life_insurance_risk'), py.get('life_insurance_risk'))}
Operational Risk:     {fmt(cy.get('op_risk'))}  {fmt(py.get('op_risk'))}  {pchg(cy.get('op_risk'), py.get('op_risk'))}
MOCE:                 {fmt(cy.get('moce'))}  {fmt(py.get('moce'))}  {pchg(cy.get('moce'), py.get('moce'))}

Write 3-4 sentences as a senior actuary commenting on the capital position.
Focus on: (1) whether the insurer remains well-capitalised, (2) key drivers of PCA change,
(3) any notable trends or concerns. Professional tone. Be specific with numbers.
"""
    url = BASE_URL.rstrip("/") + "/chat/completions"
    headers = {"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"}
    payload = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": "You are a senior actuary writing HKIA capital adequacy commentary."},
            {"role": "user",   "content": msg},
        ],
        "max_tokens": 300, "stream": True,
    }
    collected = []
    try:
        with httpx.Client(verify=False, timeout=90) as client:
            with client.stream("POST", url, headers=headers, json=payload) as resp:
                for line in resp.iter_lines():
                    line = line.strip()
                    if not line or not line.startswith("data:"): continue
                    s = line[len("data:"):].strip()
                    if s == "[DONE]": break
                    try:
                        content = json.loads(s)["choices"][0]["delta"].get("content") or ""
                        if content: collected.append(content)
                    except Exception: continue
        return "".join(collected).strip()
    except Exception as e:
        return f"[Capital commentary unavailable: {e}]"


def extract_hkrbc_data(cy_filepath: str, py_filepath: str) -> dict:
    cy = _extract_single_file(cy_filepath)
    py = _extract_single_file(py_filepath)
    return {"cy": cy, "py": py, "ai_commentary": _generate_capital_commentary(cy, py)}