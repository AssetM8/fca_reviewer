"""
debug_ca1.py  —  scans ALL columns to find where labels actually live
Run with: py debug_ca1.py "path/to/your/file.xlsx"
"""
import sys
import openpyxl

if len(sys.argv) < 2:
    print("Usage: py debug_ca1.py <path_to_excel_file>")
    sys.exit(1)

wb = openpyxl.load_workbook(sys.argv[1], data_only=True)

ca1_ws = None
for name in wb.sheetnames:
    n = name.upper()
    if "CA.1" in n and "CA.R" not in n and "CA.P" not in n:
        ca1_ws = wb[name]
        print(f"Using sheet: {name!r}\n")
        break

if not ca1_ws:
    print("CA.1 sheet not found"); sys.exit(1)

# Print every non-empty row showing ALL columns (up to col 8)
print(f"{'Row':>4}  " + "  ".join(f"Col{i:<9}" for i in range(8)))
print("-" * 100)

for i, row in enumerate(ca1_ws.iter_rows(values_only=True)):
    # Skip rows where everything is None
    if all(c is None for c in row):
        continue
    # Only print rows that have at least one non-None, non-empty value
    vals = [str(c)[:14] if c is not None else "" for c in list(row)[:8]]
    if any(v.strip() for v in vals):
        print(f"{i:>4}  " + "  ".join(f"{v:<14}" for v in vals))